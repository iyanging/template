import base64
from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import Any, Callable, Optional

from app import types
from app.redis import redis
from app.tasks import begin_meeting
from openpyxl import Workbook


async def fetch_many(
    records: list,
    record_key: str,
    record_item_attr: str,
    loader,
    item_handler: Callable[[Any], None] = None,
):
    def empty_handler(x):  # noqa
        pass

    if item_handler is None:
        item_handler = empty_handler

    items = await loader.load_many([getattr(r, record_key) for r in records])
    for record, record_items in zip(records, items):
        if not record_items:
            continue
        for item in record_items:
            item_handler(item)
        setattr(record, record_item_attr, record_items)


async def begin_meeting_task(meeting_id: types.ID, begin_at: datetime):
    task = begin_meeting.apply_async((meeting_id,), eta=begin_at)
    await redis.execute('SET', f'start_inspection_task_{meeting_id}', task.id)


def datetime_to_ymd(date: Optional[datetime], default='') -> str:
    if date:
        return date.strftime('%Y-%m-%d')
    return default


def get_attr_by_path(obj, path):
    if path == "self":
        return obj
    nodes = path.split('.') if type(path) == str else path
    node = nodes[0]
    if hasattr(obj, node):
        if len(nodes) > 1:
            return get_attr_by_path(getattr(obj, node), nodes[1:])
        else:
            return getattr(obj, node)
    else:
        return ""


def add_excel_prefix(content):
    return (
        f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{content}'
    )


def export_excel(records, header, cols) -> str:
    wb = Workbook()
    ws = wb.active
    row_no = 1
    col_no = 1
    for h in header:
        ws.cell(row_no, col_no, value=h)
        col_no += 1
    row_no += 1
    for r in records:
        col_no = 1
        for c in cols:
            key = c[0]
            function = None if len(c) < 2 else c[1]
            value = get_attr_by_path(r, key)
            if value and function:
                value = function(value)
            ws.cell(row_no, col_no, value=value)
            col_no += 1
        row_no += 1
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        content = base64.b64encode(tmp.read()).decode()
    return add_excel_prefix(content)
